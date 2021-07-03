class GetInfo:
    # seq. date. no1 ... 6 bns
    def getLottoWinInfo(self,startN, endN):
        import requests
        ret=[]
        for i in range(startN, endN+1, 1):
            req_url = "https://www.dhlottery.co.kr/common.do?method=getLottoNumber&drwNo=" + str(i)
            req = requests.get(req_url)
            if req.json()['returnValue'] == 'success':
                print("{}: success".format(str(i)))
                ret.append([str(req.json()['drwNo']), str(req.json()['drwNoDate']), str(req.json()['drwtNo1']), str(req.json()['drwtNo2']), str(req.json()['drwtNo3']), str(req.json()['drwtNo4']), str(req.json()['drwtNo5']), str(req.json()['drwtNo6']), str(req.json()['bnusNo'])])
            else:
                print("{}: fail".format(str(i)))
        return ret

    def latestSatNum(self):
        import datetime
        dn=0
        w=0
        today = datetime.datetime.today()
        while w < 6:
          tday = today - datetime.timedelta(days=dn)
          a=int(str(tday.year)[0:2])
          b=int(str(tday.year)[2:5])
          c=int(str(tday.month))
          if c <= 2:
              b = b - 1
              c = c + 12
          d=int(str(tday.day))
          w=int((((21*a/4)+(5*b/4)+(26*(c+1)/10)+d-1)%7))
          dn = dn + 1

        wSeq = 0
        sdate = datetime.date(2002,12,7)
        while sdate.year != tday.year or sdate.month != tday.month or sdate.day != tday.day:
          weak = datetime.timedelta(weeks=wSeq)
          sdate = datetime.date(2002,12,7) + weak
          if today.year == sdate.year and today.month == sdate.month and today.day == sdate.day and today.hour < 20 and today.minute < 50: break
          wSeq = wSeq + 1

        e_Seq = wSeq
        return e_Seq
        #####################################################################

    def saveInfoToHTML(self):
        info=self.getLottoWinInfo(self.startN, self.endN)
        fp = open(self.HtmlFileName, 'w')
        fp.write(  '<html>\n  <head>\n    <title>MyList</title>\n  </head>\n')
        fp.write('  <body>\n')
        fp.write('    <table border="2" align="left" width="100%" cellpadding="5" cellspacing="0">\n')
        fp.write('      <tr id="fieldname">\n')
        fp.write('        <th>Seq</th>\n')
        fp.write('        <th>Date</th>\n')
        fp.write('        <th>Num1</th>\n')
        fp.write('        <th>Num2</th>\n')
        fp.write('        <th>Num3</th>\n')
        fp.write('        <th>Num4</th>\n')
        fp.write('        <th>Num5</th>\n')
        fp.write('        <th>Num6</th>\n')
        fp.write('        <th>Bonus</th>\n')
        fp.write('      </tr>\n' )
        for i in range(len(info)):
            fp.write("      <tr id=\"seq_{}\">\n".format(info[i][0]))
            fp.write("        <th>{}</th>\n".format(info[i][0]))
            fp.write("        <th>{}</th>\n".format(info[i][1]))
            fp.write("        <th>{}</th>\n".format(info[i][2]))
            fp.write("        <th>{}</th>\n".format(info[i][3]))
            fp.write("        <th>{}</th>\n".format(info[i][4]))
            fp.write("        <th>{}</th>\n".format(info[i][5]))
            fp.write("        <th>{}</th>\n".format(info[i][6]))
            fp.write("        <th>{}</th>\n".format(info[i][7]))
            fp.write("        <th>{}</th>\n".format(info[i][8]))
            fp.write("      </tr>\n")
        fp.write('    </table>\n  </body>\n</html>')
        fp.close()
    
    def loadInfoFromHTML(self):
        from bs4 import BeautifulSoup
        with open(self.HtmlFileName, 'r') as html_fp:
            webpage =  html_fp.read()
        soup = BeautifulSoup(webpage,features="lxml")
        trList = soup.find_all('tr')
        ret=[]
        for i in range(1,len(trList)):
            ret.append([trList[i].find_all('th')[2].text, trList[i].find_all('th')[3].text, trList[i].find_all('th')[4].text, trList[i].find_all('th')[5].text, trList[i].find_all('th')[6].text, trList[i].find_all('th')[7].text, trList[i].find_all('th')[8].text])
        return ret

    def getInfoFromExcel(self):
        from openpyxl import Workbook
        from openpyxl.styles import Font, Alignment
        from openpyxl.styles import Border, Side
        from openpyxl.styles import PatternFill, Color
        from openpyxl.styles import NamedStyle
        from openpyxl.comments import Comment
        from openpyxl import load_workbook
        info = []
        wb = load_workbook(self.ExcelFileName)
        ws=wb[wb.sheetnames[0]]
        i = 2
        j = 3
        while ws.cell(row=i, column=j).value is not None:
            tmp = []
            while ws.cell(row=i, column=j).value is not None:
                #######################################
                # TO DO:
                #######################################
                if j == 3 or j == 4 or j == 5 or j == 6 or j == 7 or j == 8 or j == 9:
                    tmp.append(ws.cell(row=i, column=j).value)
                    j += 1
                else:
                    break
                #######################################
            if len(tmp) == 7:
                info.append([tmp[0],tmp[1],tmp[2],tmp[3],tmp[4],tmp[5],tmp[6]])
                tmp.clear()
            i += 1
            j = 3
        wb.close()
        return info

    def __del__(self):
        pass

    def __new__(self,IsGetAllInfo):
        return super().__new__(self)

    def __init__(self,IsGetAllInfo):
        self.startN=1
        self.endN= self.latestSatNum()
        self.ExcelFileName = './{}.xlsx'.format('Lotto_List')
        self.HtmlFileName = './{}.html'.format('MyList')
        #   file:///share/myproj/MyList.html
        if IsGetAllInfo == 1: self.saveInfoToHTML()
        self.info = self.loadInfoFromHTML()
        super().__init__()
    
    def getInfo(self):
        return self.info

if __name__ == '__main__':
    print('This is getinfo.py, Do execute main.py')
