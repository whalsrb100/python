### 0. 각 자리수별로 따로 수행.
### 1. 확인하고자하는 회차의 이전 N 개 회차의 출현 수를 확인하고
### 2. 이전 회차의 당첨번호가 몇개 회차 이전에 출현한 수인지 확인(이전 N 개 회차를 모두) 하여 리스트 작성
### 3. 이전에 생성한 리스트를 참고하여 몇개 회차 이전에 출현한 수를 선택할것인지 결정 ( 자리수당 몇개의 숫자를 추출 할것인지 고민이 필요함 )
class Myproj:
    #def __new__(self):
    #    print('new')
    #    return super().__new__(self)
    def __init__(self):
        self.ExcelFileName = './{}.xlsx'.format('Lotto_List')
        self.e_Seq = self.getLastSequence()
        self.info = self.getInfoFromExcel()
        # info[seq][0:7] : n1=0, n2=1...bns=6.

        self.ckRange = 256
        self.maxSetNum = 7
        self.StartSeq = 0
        self.tmp = self.createNums(self.ckRange)
        self.futureList = []
        for i in range(len(self.tmp)):
            self.futureList.append(list(set(self.tmp[i])))
            self.futureList[i].sort()
#######################################################
        for i in range ( 257,len(self.info)):
            self.r2 = self.createIdxLst(i)
            print("seq: {}\t{}".format(i+1,self.r2 ))
#######################################################
        super().__init__()
        exit()

        #self.checkDiff(self.futureList)

        self.CheckSeq = len(self.info)-3
        print('################### Start: createNumForOneSeq ###################')
        self.modList = self.createNumForOneSeq(self.CheckSeq)
        print('################### Finished: createNumForOneSeq ###################')
        #print('################### self.modList')
        #print(self.modList)

        self.modNum=len(self.modList)-1
        #for i in range(len(self.modList)):
        #    print("{}/{})\t{}".format(i,self.modNum,self.modList[i]))
        #exit()
        print('################### Start: checkDiff_s ###################')
        #for i in range(0,self.modNum+1):
        #    self.checkDiff_s(self.CheckSeq,self.modList[i])
        #964
        # 0 1   :   0.426416
        # 1 2   :   0.535168
        # 2 3   :   0.30104
        # 3 4   :   0.410083
        # 4 5   :   0.426393
        # 5 6   :   0.448185
        # 6 7   :   0.404637
        # 7 8   :   7.63653
        # 8 9   :   4.81317
        # 9 10  :   0.399169
        self.Range=7
        for i in range(int(self.modNum/10*self.Range),int(self.modNum/10*(self.Range+1))): self.checkDiff_s(self.CheckSeq,self.modList[i])

        #for i in range(int(self.modNum/10*6),int(self.modNum/10*7)): self.checkDiff_s(self.CheckSeq,self.modList[i])
        print('################### Finished: checkDiff_s ###################')
        super().__init__()


    def getLastSequence(self):
        import datetime
        dn=0
        w=0
        today = datetime.datetime.today()
        while w < 6:
            tday = today - datetime.timedelta(days=dn)
            a=int(str(tday.year)[0:2])
            b=int(str(tday.year)[2:5])
            c=int(str(tday.month))
            if c <= 2:
                b = b - 1
                c = c + 12
            d=int(str(tday.day))
            w=int((((21*a/4)+(5*b/4)+(26*(c+1)/10)+d-1)%7))
            dn = dn + 1
        wSeq = 0
        sdate = datetime.date(2002,12,7)
        while sdate.year != tday.year or sdate.month != tday.month or sdate.day != tday.day:
            weak = datetime.timedelta(weeks=wSeq)
            sdate = datetime.date(2002,12,7) + weak
            if today.year == sdate.year and today.month == sdate.month and today.day == sdate.day and today.hour < 20 and today.minute < 50: break
            wSeq = wSeq + 1
        return wSeq

    def getInfoFromExcel(self):
        from openpyxl import Workbook
        from openpyxl.styles import Font, Alignment
        from openpyxl.styles import Border, Side
        from openpyxl.styles import PatternFill, Color
        from openpyxl.styles import NamedStyle
        from openpyxl.comments import Comment
        from openpyxl import load_workbook
        info = []
        wb = load_workbook(self.ExcelFileName)
        ws=wb[wb.sheetnames[0]]
        i = 2
        j = 3
        while ws.cell(row=i, column=j).value is not None:
            tmp = []
            while ws.cell(row=i, column=j).value is not None:
                #######################################
                # TO DO:
                #######################################
                if j == 3 or j == 4 or j == 5 or j == 6 or j == 7 or j == 8 or j == 9:
                    tmp.append(ws.cell(row=i, column=j).value)
                    j += 1
                else:
                    break
                #######################################
            if len(tmp) == 7:
                info.append([tmp[0],tmp[1],tmp[2],tmp[3],tmp[4],tmp[5],tmp[6]])
                tmp.clear()
            i += 1
            j = 3
        wb.close()
        # info[seq][0:7] : n1=0, n2=1...bns=6.
        return info


    def createNumForOneSeq(self,seq):
        if seq >= self.e_Seq:
            print("SEQ ERROR: GREATER THEN LAST SEQ!!")
            exit(1)
        elif seq <= self.StartSeq:
            print("SEQ ERROR: LESSEST THEN FIRST SEQ!!")
            exit(1)
        oneSeqFuture = []
        for n1 in range(1,len(self.futureList[seq-self.StartSeq])-5):
            for n2 in range(n1+1,len(self.futureList[seq-self.StartSeq])-4):
                if n2 <= n1: continue
                for n3 in range(n2+1,len(self.futureList[seq-self.StartSeq])-3):
                    if n3 <= n2: continue
                    for n4 in range(n3+1,len(self.futureList[seq-self.StartSeq])-2):
                        if n4 <= n3: continue
                        for n5 in range(n4+1,len(self.futureList[seq-self.StartSeq])-1):
                            if n5 <= n4: continue
                            for n6 in range(n5+1,len(self.futureList[seq-self.StartSeq])-0):
                                if n6 <= n5: continue
                                sumTotal = self.futureList[seq-self.StartSeq][n1]+self.futureList[seq-self.StartSeq][n2]+self.futureList[seq-self.StartSeq][n3]+self.futureList[seq-self.StartSeq][n4]+self.futureList[seq-self.StartSeq][n5]+self.futureList[seq-self.StartSeq][n6]
                                #sumTotal=0
                                # Min: 21   ( 01 ~ 06 )
                                # max: 215  ( 41 ~ 45 )
                                limitMin = 121
                                limitMax = 160
                                if sumTotal > limitMin and sumTotal < limitMax:
                                    #print('##########################################################################')
                                    #print(self.futureList[seq-self.StartSeq])
                                    #print('##########################################################################')
                                    checkN=(self.futureList[seq-self.StartSeq][n1]%2)+(self.futureList[seq-self.StartSeq][n2]%2)+(self.futureList[seq-self.StartSeq][n3]%2)+(self.futureList[seq-self.StartSeq][n4]%2)+(self.futureList[seq-self.StartSeq][n5]%2)+(self.futureList[seq-self.StartSeq][n6]%2)
                                    if checkN <= 1 or checkN >= 5: continue
                                    if self.futureList[seq-self.StartSeq][n1] >= 10: continue
                                    if self.futureList[seq-self.StartSeq][n2] >= 35: continue
                                    if (self.futureList[seq-self.StartSeq][n2]-self.futureList[seq-self.StartSeq][n1] <= 5): continue
                                    if self.futureList[seq-self.StartSeq][n3]-self.futureList[seq-self.StartSeq][n2] <= 1:
                                        if self.futureList[seq-self.StartSeq][n4]-self.futureList[seq-self.StartSeq][n3] <= 1:
                                            if self.futureList[seq-self.StartSeq][n5]-self.futureList[seq-self.StartSeq][n4] <= 1 : continue
                                    if (self.futureList[seq-self.StartSeq][n6]-self.futureList[seq-self.StartSeq][n5] <= 5): continue

                                    oneSeqFuture.append([self.futureList[seq-self.StartSeq][n1],self.futureList[seq-self.StartSeq][n2],self.futureList[seq-self.StartSeq][n3],self.futureList[seq-self.StartSeq][n4],self.futureList[seq-self.StartSeq][n5],self.futureList[seq-self.StartSeq][n6]])
        return oneSeqFuture
            
    def checkDiff_s(self,seq,future):
        gn = 5
        cknums_s = 1
        cknums_e = 6
        TotalCheckNum = 0
        self.grade = []
        for i in range(0,5):
            self.grade.append(0)
        for i in range(seq,seq+1):
            checkNum = 0
            isLast = False
            for j in range(cknums_s-1,cknums_e):
                ckDup = []
                try:
                    if self.info[i][j] in future:      checkNum += 1
                except IndexError: isLast = True
                ckDup.clear()
            check2nd=0
            if isLast: break
            if self.info[i][6] in future:         check2nd=1
            if checkNum >= 3:
                TotalCheckNum += 1
                if    checkNum == 6: self.grade[4] += 1
                elif checkNum == 5:
                    if check2nd > 0: self.grade[3] += 1
                    else: self.grade[2] += 1
                elif checkNum == 4: self.grade[1] += 1
                elif checkNum == 3: self.grade[0] += 1
            if   checkNum == 6:                                  print("\t!!!1) seq({})\t{}\t:\t({})\t{}".format(i+1,self.info[i][0:6], len(future),future).replace('[','').replace(']',''))
            elif checkNum == 5 and check2nd == 1:  print("\t@2) seq({})\t{}\t:\t({})\t{}".format(i+1,self.info[i][0:6], len(future),future).replace('[','').replace(']',''))
            elif checkNum == 5:                                  print("\t#3) seq({})\t{}\t:\t({})\t{}".format(i+1,self.info[i][0:6], len(future),future).replace('[','').replace(']',''))
            elif checkNum == 4:                                  print("\t$4) seq({})\t{}\t:\t({})\t{}".format(i+1,self.info[i][0:6], len(future),future).replace('[','').replace(']',''))
            elif checkNum == 3:                                  print("\t%5) seq({})\t{}\t:\t({})\t{}".format(i+1,self.info[i][0:6], len(future),future).replace('[','').replace(']',''))
            else:                                                            print("\tX) seq({})\t{}\t:\t({})\t{}".format(i+1,self.info[i][0:6], len(future),future).replace('[','').replace(']',''))
        if TotalCheckNum > 0 : print("({}/{}*100):{:0.6f}%\t({}/{}) 5:{}, 4:{}, 3:{}, 2:{}, 1:{}".format(TotalCheckNum, len(self.modList), TotalCheckNum/len(self.modList)*100.0, TotalCheckNum,len(self.modList),self.grade[0],self.grade[1],self.grade[2],self.grade[3],self.grade[4]))
        

    def checkDiff(self,future):
        gn = 5
        cknums_s = 1
        cknums_e = 6
        TotalCheckNum = 0
        self.grade = []
        for i in range(0,5): self.grade.append(0)
        for i in range(self.StartSeq, len(self.info)):
            checkNum = 0
            isLast = False
            for j in range(cknums_s-1,cknums_e):
                ckDup = []
                #print(    "futures Num:{}, {}".format(    len(self.future[i-self.StartSeq]), list(set(self.future[i-self.StartSeq]))    )    )
                try:
                    if self.info[i][j] in future[i-self.StartSeq]:      checkNum += 1
                except IndexError: isLast = True
                ckDup.clear()
            check2nd=0
            if isLast: break
            if self.info[i][6] in future[i-self.StartSeq]:         check2nd=1
            #print("{}: n{}\t{} ::::: {} ".format( i, checkNum, self.info[i][0:6],self.future[i-(self.StartSeq)]  ) )
            if checkNum >= 3:
                #if checkNum >= 3:
                    #if i > 900: print("{}: n{}\t{} {} {} {} {} {} ::::: {} {} {} {} {} {} ".format( i, checkNum, self.info[i][0],self.info[i][1],self.info[i][2],self.info[i][3],self.info[i][4],self.info[i][5],self.future[i-self.StartSeq][0],self.future[i-self.StartSeq][1],self.future[i-self.StartSeq][2],self.future[i-self.StartSeq][3],self.future[i-self.StartSeq][4],self.future[i-self.StartSeq][5]  ) )
                #print("{}:{}개".format(i+1,checkNum))
                TotalCheckNum += 1
                if    checkNum == 6: self.grade[4] += 1
                elif checkNum == 5:
                    if check2nd > 0: self.grade[3] += 1
                    else: self.grade[2] += 1
                elif checkNum == 4: self.grade[1] += 1
                elif checkNum == 3: self.grade[0] += 1

            if   checkNum == 6:                                  print("1) seq({})\t{}\t:\t({})\t{}".format(i+1,self.info[i][0:6], len(future[i-self.StartSeq]),future[i-self.StartSeq]).replace('[','').replace(']',''))
            elif checkNum == 5 and check2nd == 1:  print("2) seq({})\t{}\t:\t({})\t{}".format(i+1,self.info[i][0:6], len(future[i-self.StartSeq]),future[i-self.StartSeq]).replace('[','').replace(']',''))
            elif checkNum == 5:                                  print("3) seq({})\t{}\t:\t({})\t{}".format(i+1,self.info[i][0:6], len(future[i-self.StartSeq]),future[i-self.StartSeq]).replace('[','').replace(']',''))
            elif checkNum == 4:                                  print("4) seq({})\t{}\t:\t({})\t{}".format(i+1,self.info[i][0:6], len(future[i-self.StartSeq]),future[i-self.StartSeq]).replace('[','').replace(']',''))
            elif checkNum == 3:                                  print("5) seq({})\t{}\t:\t({})\t{}".format(i+1,self.info[i][0:6], len(future[i-self.StartSeq]),future[i-self.StartSeq]).replace('[','').replace(']',''))
            else:                                                            print("X) seq({})\t{}\t:\t({})\t{}".format(i+1,self.info[i][0:6], len(future[i-self.StartSeq]),future[i-self.StartSeq]).replace('[','').replace(']',''))
        print("*) seq({})\t*** Next ***\t\t:\t({})\t{} ".format(i+2,len(future[-1]),future[-1]).replace('[','').replace(']',''))
        print("({}/{}*100):{:0.6f}%\t({}/{}) 5:{}, 4:{}, 3:{}, 2:{}, 1:{}".format(TotalCheckNum, (len(self.info)-self.ckRange), TotalCheckNum/(len(self.info)-self.ckRange)*100.0, TotalCheckNum,len(self.info)-self.ckRange,self.grade[0],self.grade[1],self.grade[2],self.grade[3],self.grade[4]))


    def cleanUp(self):
        for i in range(self.s_Seq - self.s_Seq, self.e_Seq - self.s_Seq):
            if self.info[i]: self.info[i].clear()
        if self.future: self.future.clear()
        for i in range(0,len(self.future)): self.future[i].clear()
        if self.future: self.future.clear()
        if self.grade: self.grade.clear()

    def getValidHistoryList(self, n, info):
        validHistoryList = []
        for i in range(n,n+1):
            validHistoryDict = {}
            for j in range(0, len(info)):
                try: validHistoryDict[info[j][i]] += 1
                except KeyError: validHistoryDict[info[j][i]] = 1

        #keysList = list(validHistoryDict)
        #keysList = list(dict(  sorted(validHistoryDict.items(), key=(lambda x: x[0]) ,reverse=False)))
        if    n == 0: keysList = list(dict(  sorted(validHistoryDict.items(), key=(lambda x: x[0]) ,reverse=True)))
        elif n == 1: keysList = list(dict(  sorted(validHistoryDict.items(), key=(lambda x: x[0]) ,reverse=True)))
        elif n == 2: keysList = list(dict(  sorted(validHistoryDict.items(), key=(lambda x: x[0]) ,reverse=True)))
        elif n == 3: keysList = list(dict(  sorted(validHistoryDict.items(), key=(lambda x: x[0]) ,reverse=False)))
        elif n == 4: keysList = list(dict(  sorted(validHistoryDict.items(), key=(lambda x: x[0]) ,reverse=False)))
        elif n == 5: keysList = list(dict(  sorted(validHistoryDict.items(), key=(lambda x: x[0]) ,reverse=False)))

        for i in range(0,len(keysList)):
            validHistoryList.append(keysList[i])
            validHistoryList.append(validHistoryDict[keysList[i]])
            validHistoryList.append(0)
        return validHistoryList


    def getValidHistoryList2(self, n, info):
        validHistoryList = []
        for i in range(n,n+1):
            for j in range(0, len(info)):
                validHistoryList.append(info[j][i])
        return validHistoryList

    def createIdxLst(self,num):
        history=[]
        ret = []
        IsDebug=False
        if num - self.StartSeq < 1: StartNum = num
        else: StartNum = self.StartSeq
        for i in range(num, num+1):
            for j in range(0,1):
                history.append(self.getValidHistoryList2(j,self.info[i-(StartNum):i]))
                if IsDebug:
                    #print("\t########################## j: {} ##########################".format(j))
                    print("\t({}/6) info: {}".format(j+1,self.info[i][j]))
                    print("\t({}/6) history: {}".format(j+1,history[j]))
                    #print("\t##########################################################")
                cnt=0
                for k in reversed(range(  int(len(history[j]) )  )):
                    cnt += 1
                    if self.info[i][j] == history[j][k]:
                        ret.append(cnt)
                        break
                
        return ret


    def createNums(self,ckRange):
        nums = []    # return nums
        history=[]
        setNum=0
        # ckRange : 한번 번호 추출할때 참조할 회차의 개수 (default: 100)
        self.StartSeq =  len(self.info) - 1
        if len(self.info) > ckRange:            self.StartSeq =  ckRange

        for i in range(self.StartSeq, len(self.info)+1):
            nums.append([])
            for j in range(0,6):
                nums[i-(self.StartSeq)].append([])
                for k in range(self.maxSetNum): nums[i-(self.StartSeq)][j].append(0)
            for j in range(0,6):
                history.append(self.getValidHistoryList(j,self.info[i-(self.StartSeq):i-1]))
                for k in range (  int(len(history[j]) / 3)  ): history[j][k*3+2] = 0
            while True:
                for j in range(0,6):
                    modtime=0.90
                    IsBreake = False
                    while IsBreake == False:
                        for k in range (  int(len(history[j]) / 3)  ):
                            if history[j][k*3+1] >= int((ckRange / int(len(history[j]) / 3)) * modtime): history[j][k*3+2] += 1
                        for k in range (  int(len(history[j]) / 3)  ):
                            if (history[j][k*3+2] + history[j][k*3+1] ) % 100 == 0:
                                if nums[i-(self.StartSeq)][j][setNum] == 0: nums[i-(self.StartSeq)][j][setNum] = history[j][k*3+0]
                                setNum += 1
                                if setNum == self.maxSetNum:
                                    setNum=0
                                    break
                        IsBreake = True
                        for k in range(0,self.maxSetNum):
                            if nums[i-(self.StartSeq)][j][k] == 0: IsBreake = False
                if IsBreake: break
            history.clear()
            ret = []
            for i in range(len(nums)):
                ret.append([])
                for j in range(len(nums[i])):
                    for k in range(len(nums[i][j])):
                        ret[i].append(nums[i][j][k])
        nums.clear()
        return ret



    def __del__(self):
        #self.cleanUp()
        pass

def main():
    import sys
    f=Myproj()
    del f


if __name__ == '__main__':
    main()
