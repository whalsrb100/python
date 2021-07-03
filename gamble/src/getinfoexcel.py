from openpyxl import Workbook
from openpyxl.styles import Font, Alignment
from openpyxl.styles import Border, Side
from openpyxl.styles import PatternFill, Color
from openpyxl.styles import NamedStyle
from openpyxl.comments import Comment
from openpyxl import load_workbook
#회차	추첨일	Num1	Num2	Num3	Num4	Num5	Num6	Bonus
#1	2002-12-07	10	23	29	33	37	40	16


def getInfoExcel(Excel_FileName):
    info = []
    wb = load_workbook(Excel_FileName)
    ws=wb[wb.sheetnames[0]]
    i = 2
    j = 3
    while ws.cell(row=i, column=j).value is not None:
        tmp = []
        while ws.cell(row=i, column=j).value is not None:
            #######################################
            # TO DO:
            #######################################
            if j == 3 or j == 4 or j == 5 or j == 6 or j == 7 or j == 8 or j == 9:
                tmp.append(ws.cell(row=i, column=j).value)
                j += 1
            else:
                break
            #######################################

        if len(tmp) == 7:
            info.append([tmp[0],tmp[1],tmp[2],tmp[3],tmp[4],tmp[5],tmp[6]])
            tmp.clear()
        i += 1
        j = 3

    #wb.save(Excel_FileName)
    wb.close()
    return info
