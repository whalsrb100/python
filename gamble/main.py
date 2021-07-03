import sys
#import requests
#from tqdm import tqdm
#import json
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment
from openpyxl.styles import Border, Side
from openpyxl.styles import PatternFill, Color
from openpyxl.styles import NamedStyle
from openpyxl.comments import Comment
from openpyxl import load_workbook
#!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!#
# excel 파일을 읽어 내용을 수정
#!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!#

###########################################################################
# Set seq
from src.getlastseq import getLastSequence
###########################################################################
s_Seq = 1   # 시작 회차
e_Seq = getLastSequence() # 마지막 회차
###########################################################################


###########################################################################
# excel 파일을 읽어 정보를 수집
from src.getinfoexcel import getInfoExcel
###########################################################################
Excel_FileName = './' + 'Lotto_List' + '.xlsx'
info = getInfoExcel(Excel_FileName)

#    N1        Num3    Num4    Num5    Num6    Bonus
#    10    23    29    33    37    40    16
#e_Seq = 2
indexPool = []
for i in range(s_Seq - s_Seq, e_Seq - s_Seq):
    if i > 0:
        preinfo = info[i-1]
    indexPool.append((40 * 41 * 42 * 43 * 44 * info[i][0]) + (40 * 41 * 42 * 43  * info[i][1]) + (40 * 41 * 42 * info[i][2]) + (40 * 41 * info[i][3]) + (40 * info[i][4]) + (info[i][5]))
    print("{}:\t{}\t{}\t{}\t{}\t{}\t{}\tb:{}\t{}".format(str(i+1),info[i][0],info[i][1],info[i][2],info[i][3],info[i][4],info[i][5],info[i][6],indexPool[i]))
###########################################################################

###########################################################################
# create nums
###########################################################################
from src.createnums import createNums
if len(info) > 100:
    max_N = 100
else:
    max_N = len(info) - 1
N = max_N
future = []
for i in range(N, len(info)):
    future.append(createNums(info[i-N:i],indexPool[i-1]))

#print(future)
###########################################################################

###########################################################################
# Diff
###########################################################################
TotalCheckNum = 0
grade = []
for i in range(0,4): grade.append(0)
for i in range(N, len(info)):
    checkNum = 0
    for j in range(0,6):
        ckDup = []
        for k in range(0,6):
            if info[i][j] == future[i-N][k]:
                if future[i-N][k] not in ckDup:
                    checkNum += 1
                    ckDup.append(future[i-N][k])
        ckDup.clear()
    if checkNum >= 3:
        if checkNum >= 3: print("n{}\t{} {} {} {} {} {} ::::: {} {} {} {} {} {} ".format( checkNum, info[i][0],info[i][1],info[i][2],info[i][3],info[i][4],info[i][5],future[i-N][0],future[i-N][1],future[i-N][2],future[i-N][3],future[i-N][4],future[i-N][5]  ) )
        #print("{}:{}개".format(i+1,checkNum))
        TotalCheckNum += 1
        if    checkNum == 6: grade[3] += 1
        elif checkNum == 5: grade[2] += 1
        elif checkNum == 4: grade[1] += 1
        elif checkNum == 3: grade[0] += 1
print("{}%\t({}/{}) 5:{}, 4:{}, 3:{}, 1:{}".format(TotalCheckNum/(len(info))*100.0, TotalCheckNum,len(info)-1,grade[0],grade[1],grade[2],grade[3]))
###########################################################################

###########################################################################
# Clean up
###########################################################################
indexPool.clear()
for i in range(s_Seq - s_Seq, e_Seq - s_Seq): info[i].clear()
info.clear()
for i in range(0,len(future)): future[i].clear()
future.clear()
grade.clear()
#print("Clearn up")
###########################################################################
exit()


###########################################################################
# Open File : Excel
###########################################################################
#Excel_FileName = 'Lotto_List' + '.xlsx'
Excel_FileName = '/share/myproj/python/lot/' + 'Lotto_List' + '.xlsx'
wb = load_workbook(Excel_FileName)
ws=wb[wb.sheetnames[0]]

myarr = []
myque = []
for i in range(8): myque.append([])

cnt=0
s_add=0
for i in range(s_Seq - s_Seq , e_Seq - s_Seq + 1):
    cnt = cnt+1
    scale = 0.8
    scale_help = 0.000001
    index_init = int(ws['S'+str(i+2)].value)
    indexn = index_init * scale
    myarr.append ( int(indexn/130320960)+1 )
    indexn = indexn % 130320960
    myarr.append ( int(indexn/2961840)+2 )
    indexn = indexn % 2961840
    myarr.append ( int(indexn/68880)+3 )
    indexn = indexn % 68880
    myarr.append ( int(indexn/1640)+4 )
    indexn = indexn % 1640
    myarr.append ( int(indexn/40)+5 )
    indexn = indexn % 40
    myarr.append ( int(indexn)+6 )
    #====================================================#
    if myarr[0] < 1: myarr[0] = 1
    while myarr[0] >= myarr[1] or myarr[1] >= myarr[2] or myarr[2] >= myarr[3] or myarr[3] >= myarr[4] or myarr[4] >= myarr[5]:
        if myarr[0] >= myarr[1]: myarr[1] += 1
        if myarr[1] >= myarr[2]: myarr[2] += 1
        if myarr[2] >= myarr[3]: myarr[3] += 1
        if myarr[3] >= myarr[4]: myarr[4] += 1
        if myarr[4] >= myarr[5]: myarr[5] += 1

    myarrsum = myarr[0]+myarr[1]+myarr[2]+myarr[3]+myarr[4]+myarr[5]
    mymin = 60
    mymax = 95
    scnt = 0
    while myarrsum > mymax or myarrsum < mymin:
        scnt += 1
        if scnt > 10000:
            scnt = 0
            scale_help *= 0.1
            mymax += 1
        myarr.clear()
        if myarrsum < mymin: scale += scale_help
        elif myarrsum > mymax: scale -= scale_help
        indexn = index_init * scale
        myarr.append ( int(indexn/130320960)+1 )
        indexn = indexn % 130320960
        myarr.append ( int(indexn/2961840)+2 )
        indexn = indexn % 2961840
        myarr.append ( int(indexn/68880)+3 )
        indexn = indexn % 68880
        myarr.append ( int(indexn/1640)+4 )
        indexn = indexn % 1640
        myarr.append ( int(indexn/40)+5 )
        indexn = indexn % 40
        myarr.append ( int(indexn)+6 )
        myarrsum = myarr[0]+myarr[1]+myarr[2]+myarr[3]+myarr[4]+myarr[5]

    if myarr[0] < 1: myarr[0] = 1
    while myarr[0] >= myarr[1] or myarr[1] >= myarr[2] or myarr[2] >= myarr[3] or myarr[3] >= myarr[4] or myarr[4] >= myarr[5]:
        if myarr[0] >= myarr[1]: myarr[1] += 1
        if myarr[1] >= myarr[2]: myarr[2] += 1
        if myarr[2] >= myarr[3]: myarr[3] += 1
        if myarr[3] >= myarr[4]: myarr[4] += 1
        if myarr[4] >= myarr[5]: myarr[5] += 1
    
    myarrsum = myarr[0]+myarr[1]+myarr[2]+myarr[3]+myarr[4]+myarr[5]
    ws['AC'+str(i+2)] = myarrsum

    ws['T'+str(i+2)] = myarr[0]
    ws['U'+str(i+2)] = myarr[1]
    ws['V'+str(i+2)] = myarr[2]
    ws['W'+str(i+2)] = myarr[3]
    ws['X'+str(i+2)] = myarr[4]
    ws['Y'+str(i+2)] = myarr[5]
  
    #############################################################
    # check diff
    #############################################################

    diffnum=[]
    if i > 0:
        myarr.clear()
        myarr.append(int(ws['T'+str(i+1)].value))
        myarr.append(int(ws['U'+str(i+1)].value))
        myarr.append(int(ws['V'+str(i+1)].value))
        myarr.append(int(ws['W'+str(i+1)].value))
        myarr.append(int(ws['X'+str(i+1)].value))
        myarr.append(int(ws['Y'+str(i+1)].value))
        diffnum.append(int(ws['C'+str(i+2)].value))
        diffnum.append(int(ws['D'+str(i+2)].value))
        diffnum.append(int(ws['E'+str(i+2)].value))
        diffnum.append(int(ws['F'+str(i+2)].value))
        diffnum.append(int(ws['G'+str(i+2)].value))
        diffnum.append(int(ws['H'+str(i+2)].value))
        diffsum = diffnum[0]+diffnum[1]+diffnum[2]+diffnum[3]+diffnum[4]+diffnum[5]
        print ("({})\t{}\t{}\t{}\t{}\t{}\t{}".format( diffsum,diffnum[0],diffnum[1],diffnum[2],diffnum[3],diffnum[4],diffnum[5] ) )
        ws['AB'+str(i+2)] = diffsum


    #====================================================#
    # queue(10) 6 nums, bnum, sum. array 8
    #====================================================#
        #viewmaxnum=10
        #for qidx in range(6):
        #    if len(myque[qidx]) < viewmaxnum:
        #        myque[qidx].append(diffnum[qidx])
        #    while len(myque[qidx]) >= viewmaxnum: myque[qidx].pop(0)
        ## bonus
        #myque[6].append(ws['I'+str(i+1)].value)
        #while len(myque[6]) >= viewmaxnum: myque[6].pop(0)
        ## sum for 6 numbers
        #myque[7].append(diffsum)
        #while len(myque[7]) >= viewmaxnum: myque[7].pop(0)
    #====================================================#

    #====================================================#
    # add Helper
    #====================================================#





    #====================================================#

        rnum = 0
        for ck_i in range(0,6):
            for ck_j in range(0,6):
              if diffnum[ck_i] == myarr[ck_j]:
                rnum = rnum + 1
    
        if rnum < 3:
            ws['Z'+str(i+1)] = "꽝 (" + str(rnum) + ")"
        elif rnum == 3:
            ws['Z'+str(i+1)] = "5등 (" + str(rnum) + ")"
            #print("myarr[0]: " + str(myarr[0]) + ", diffnum[0]: " + str(diffnum[0]) )
            #print("myarr[1]: " + str(myarr[1]) + ", diffnum[1]: " + str(diffnum[1]) )
            #print("myarr[2]: " + str(myarr[2]) + ", diffnum[2]: " + str(diffnum[2]) )
            #print("myarr[3]: " + str(myarr[3]) + ", diffnum[3]: " + str(diffnum[3]) )
            #print("myarr[4]: " + str(myarr[4]) + ", diffnum[4]: " + str(diffnum[4]) )
            #print("myarr[5]: " + str(myarr[5]) + ", diffnum[5]: " + str(diffnum[5]) )
            print(str(cnt) + " - 5등 (" + str(rnum) + ")")
        elif rnum == 4:
            ws['Z'+str(i+1)] = "4등 (" + str(rnum) + ")"
            #print("myarr[0]: " + str(myarr[0]) + ", diffnum[0]: " + str(diffnum[0]) )
            #print("myarr[1]: " + str(myarr[1]) + ", diffnum[1]: " + str(diffnum[1]) )
            #print("myarr[2]: " + str(myarr[2]) + ", diffnum[2]: " + str(diffnum[2]) )
            #print("myarr[3]: " + str(myarr[3]) + ", diffnum[3]: " + str(diffnum[3]) )
            #print("myarr[4]: " + str(myarr[4]) + ", diffnum[4]: " + str(diffnum[4]) )
            #print("myarr[5]: " + str(myarr[5]) + ", diffnum[5]: " + str(diffnum[5]) )
            print(str(cnt) + " - 4등 (" + str(rnum) + ")")
        elif rnum == 5:
            ws['Z'+str(i+1)] = "3등 (" + str(rnum) + ")"
            #print("myarr[0]: " + str(myarr[0]) + ", diffnum[0]: " + str(diffnum[0]) )
            #print("myarr[1]: " + str(myarr[1]) + ", diffnum[1]: " + str(diffnum[1]) )
            #print("myarr[2]: " + str(myarr[2]) + ", diffnum[2]: " + str(diffnum[2]) )
            #print("myarr[3]: " + str(myarr[3]) + ", diffnum[3]: " + str(diffnum[3]) )
            #print("myarr[4]: " + str(myarr[4]) + ", diffnum[4]: " + str(diffnum[4]) )
            #print("myarr[5]: " + str(myarr[5]) + ", diffnum[5]: " + str(diffnum[5]) )
            print(str(cnt) + " - 3등 (" + str(rnum) + ")")
            for ck_i in range(0,6):
                if myarr[ck_i] == ws['I'+str(i+2)].value:
                    ws['Z'+str(i+1)] = "2등 (" + str(rnum) + ")"
                    #print("myarr[0]: " + str(myarr[0]) + ", diffnum[0]: " + str(diffnum[0]) )
                    #print("myarr[1]: " + str(myarr[1]) + ", diffnum[1]: " + str(diffnum[1]) )
                    #print("myarr[2]: " + str(myarr[2]) + ", diffnum[2]: " + str(diffnum[2]) )
                    #print("myarr[3]: " + str(myarr[3]) + ", diffnum[3]: " + str(diffnum[3]) )
                    #print("myarr[4]: " + str(myarr[4]) + ", diffnum[4]: " + str(diffnum[4]) )
                    #print("myarr[5]: " + str(myarr[5]) + ", diffnum[5]: " + str(diffnum[5]) )
                    print(str(cnt) + " - 2등 (" + str(rnum) + ")")
                    break
        elif rnum == 6:
            ws['Z'+str(i+1)] = "1등 (" + str(rnum) + ")"
            #print("myarr[0]: " + str(myarr[0]) + ", diffnum[0]: " + str(diffnum[0]) )
            #print("myarr[1]: " + str(myarr[1]) + ", diffnum[1]: " + str(diffnum[1]) )
            #print("myarr[2]: " + str(myarr[2]) + ", diffnum[2]: " + str(diffnum[2]) )
            #print("myarr[3]: " + str(myarr[3]) + ", diffnum[3]: " + str(diffnum[3]) )
            #print("myarr[4]: " + str(myarr[4]) + ", diffnum[4]: " + str(diffnum[4]) )
            #print("myarr[5]: " + str(myarr[5]) + ", diffnum[5]: " + str(diffnum[5]) )
            print(str(cnt) + " - 1등 (" + str(rnum) + ")")
        diffnum.clear()
    myarr.clear()
  
r1cnt = 0
r2cnt = 0
r3cnt = 0
r4cnt = 0
r5cnt = 0
##### get value for seq
for rows in wb[wb.sheetnames[0]].iter_rows():
    if str(rows[25].value) == "1등 (6)":
        r1cnt = r1cnt + 1
    elif str(rows[25].value) == "2등 (5)":
        r2cnt = r2cnt + 1
    elif str(rows[25].value) == "3등 (5)":
        r3cnt = r3cnt + 1
    elif str(rows[25].value) == "4등 (4)":
        r4cnt = r4cnt + 1
    elif str(rows[25].value) == "5등 (3)":
        r5cnt = r5cnt + 1
################################################
# Save And Close File : Excel
################################################
wb.save(Excel_FileName)
wb.close()


################################################
print('Finished [' + str(s_Seq) + '/' + str(e_Seq) + '] (' + str(r1cnt+r2cnt+r3cnt+r4cnt+r5cnt)+').')
if r1cnt > 0: print("1등:"+str(r1cnt)+"개")
if r2cnt > 0: print("2등:"+str(r2cnt)+"개")
if r3cnt > 0: print("3등:"+str(r3cnt)+"개")
if r4cnt > 0: print("4등:"+str(r4cnt)+"개")
if r5cnt > 0: print("5등:"+str(r5cnt)+"개")
