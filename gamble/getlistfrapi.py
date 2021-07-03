import pandas as pd
import requests
from tqdm import tqdm
import json
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment
from openpyxl.styles import Border, Side
from openpyxl.styles import PatternFill, Color
from openpyxl.styles import NamedStyle
from openpyxl.comments import Comment
from openpyxl import load_workbook
import datetime
import time


# pip install pandas
# pip install requests
# pip install tqdm
# pip install openpyxl

##########################################################################################################
# 1부터 현재까지 새로 가져오기
##########################################################################################################
##########################################################################################################



def getLottoWinInfo(minDrwNo, maxDrwNo):
    drwtNo1 = []
    drwtNo2 = []
    drwtNo3 = []
    drwtNo4 = []
    drwtNo5 = []
    drwtNo6 = []
    drwtSeq = []
    bnusNo = []
    totSellamnt = []
    drwNoDate = []
    firstAccumamnt = []
    firstPrzwnerCo = []
    firstWinamnt = []
    
    rawdrwtNo1 = []
    rawdrwtNo2 = []
    rawdrwtNo3 = []
    rawdrwtNo4 = []
    rawdrwtNo5 = []
    rawdrwtNo6 = []
    rawdrwtSeq = []
    rawbnusNo = []
    rawtotSellamnt = []
    rawdrwNoDate = []
    rawfirstAccumamnt = []
    rawfirstPrzwnerCo = []
    rawfirstWinamnt = []

    #for i in tqdm(range(minDrwNo, maxDrwNo+1, 1)):
    for i in range(minDrwNo, maxDrwNo+1, 1):
        req_url = "https://www.dhlottery.co.kr/common.do?method=getLottoNumber&drwNo=" + str(i)
        req_lotto = requests.get(req_url)
        print(str(i).zfill(4) + ":\t" + str(req_lotto).split('<')[1].split('>')[0])
        lottoNo = req_lotto.json()
        drwtSeq.append(str(i))
        drwtNo1.append(lottoNo['drwtNo1'])
        drwtNo2.append(lottoNo['drwtNo2'])
        drwtNo3.append(lottoNo['drwtNo3'])
        drwtNo4.append(lottoNo['drwtNo4'])
        drwtNo5.append(lottoNo['drwtNo5'])
        drwtNo6.append(lottoNo['drwtNo6'])
        bnusNo.append(lottoNo['bnusNo'])
        totSellamnt.append(lottoNo['totSellamnt'])
        drwNoDate.append(lottoNo['drwNoDate'])
        firstAccumamnt.append(lottoNo['firstAccumamnt'])
        firstPrzwnerCo.append(lottoNo['firstPrzwnerCo'])
        firstWinamnt.append(lottoNo['firstWinamnt'])
        lotto_dict = {"회차":drwtSeq, "추첨일":drwNoDate, "Num1":drwtNo1, "Num2":drwtNo2, "Num3":drwtNo3, "Num4":drwtNo4, "Num5":drwtNo5, "Num6":drwtNo6, "bnsNum":bnusNo, "총판매금액":totSellamnt, "총1등당첨금":firstAccumamnt, "1등당첨인원":firstPrzwnerCo, "1등수령액":firstWinamnt}
        #print(lotto_dict)
        df_lotto = pd.DataFrame(lotto_dict)

        rawdrwtSeq.append(str(i))
        rawdrwtNo1.append(lottoNo['drwtNo1'])
        rawdrwtNo2.append(lottoNo['drwtNo2'])
        rawdrwtNo3.append(lottoNo['drwtNo3'])
        rawdrwtNo4.append(lottoNo['drwtNo4'])
        rawdrwtNo5.append(lottoNo['drwtNo5'])
        rawdrwtNo6.append(lottoNo['drwtNo6'])
        rawbnusNo.append(lottoNo['bnusNo'])
        rawtotSellamnt.append(lottoNo['totSellamnt'])
        rawdrwNoDate.append(lottoNo['drwNoDate'])
        rawfirstAccumamnt.append(lottoNo['firstAccumamnt'])
        rawfirstPrzwnerCo.append(lottoNo['firstPrzwnerCo'])
        rawfirstWinamnt.append(lottoNo['firstWinamnt'])
    Num1=str(rawdrwtNo1[0])
    Num2=str(rawdrwtNo2[0])
    Num3=str(rawdrwtNo3[0])
    Num4=str(rawdrwtNo4[0])
    Num5=str(rawdrwtNo5[0])
    Num6=str(rawdrwtNo6[0])
    #print(Num1 + " " + Num2 + " " + Num3 + " " + Num4 + " " + Num5 + " " + Num6)
    return df_lotto
s_Seq=1
#####################################################################
# 가장 최근의 토요일 주차 구하기
#####################################################################
dn=0
w=0
today = datetime.datetime.today()
while w < 6:
  tday = today - datetime.timedelta(days=dn)
  a=int(str(tday.year)[0:2])
  b=int(str(tday.year)[2:5])
  c=int(str(tday.month))
  if c <= 2:
      b = b - 1
      c = c + 12
  d=int(str(tday.day))
  w=int((((21*a/4)+(5*b/4)+(26*(c+1)/10)+d-1)%7))
  dn = dn + 1
    
wSeq = 0
sdate = datetime.date(2002,12,7)
while sdate.year != tday.year or sdate.month != tday.month or sdate.day != tday.day:
  weak = datetime.timedelta(weeks=wSeq)
  sdate = datetime.date(2002,12,7) + weak
  if today.year == sdate.year and today.month == sdate.month and today.day == sdate.day and today.hour < 20 and today.minute < 50: break
  wSeq = wSeq + 1

e_Seq = wSeq
#####################################################################
lotto_result = getLottoWinInfo(s_Seq, e_Seq)

#print(lotto_result)

print('###################')

################################################
# Open File : Excel
################################################
#-----------------------------------------------------------------------#
font_12t = Font(name='맑은 고딕', size=12, bold=True)
font_12 = Font(name='맑은 고딕', size=12, bold=False)
align_center = Alignment(horizontal='center', vertical='center')
border_thin = Border(
    left=Side(style='thin'), right=Side(style='thin'),
    top=Side(style='thin'), bottom=Side(style='thin')
    )
fill_Yellow = PatternFill(patternType='solid',fgColor=Color('FFFF00'))
#-----------------------------------------------------------------------#

Excel_FileName = 'Lotto_List' + '.xlsx'
wb=Workbook()
ws=wb.active

ws['A'+str(1)]='회차'
ws['B'+str(1)]='추첨일'
ws['C'+str(1)]='Num1'
ws['D'+str(1)]='Num2'
ws['E'+str(1)]='Num3'
ws['F'+str(1)]='Num4'
ws['G'+str(1)]='Num5'
ws['H'+str(1)]='Num6'
ws['I'+str(1)]='Bonus'
ws['J'+str(1)]='총판매금액'
ws['K'+str(1)]='총1등당첨금'
ws['L'+str(1)]='1등당첨인원'
ws['M'+str(1)]='1등수령액'

ws['N'+str(1)]='1의자리'
ws['O'+str(1)]='10의자리'
ws['P'+str(1)]='20의자리'
ws['Q'+str(1)]='30의자리'
ws['R'+str(1)]='40의자리'
ws['S'+str(1)]='Index(Full)'

ws['T'+str(1)] = '예상1'
ws['U'+str(1)] = '예상2'
ws['V'+str(1)] = '예상3'
ws['W'+str(1)] = '예상4'
ws['X'+str(1)] = '예상5'
ws['Y'+str(1)] = '예상6'
ws['Z'+str(1)] = '예상결과'
#ws['T'+str(1)]='Index'
################################################
  
tmp = []
for i in range(s_Seq - s_Seq , e_Seq - s_Seq + 1):
#  print(lotto_result.iloc[i,0]) # sequence
#  print(lotto_result.iloc[i,1]) # date
#  print(lotto_result.iloc[i,2]) # num 1
#  print(lotto_result.iloc[i,3]) # num 2
#  print(lotto_result.iloc[i,4]) # num 3
#  print(lotto_result.iloc[i,5]) # num 4
#  print(lotto_result.iloc[i,6]) # num 5
#  print(lotto_result.iloc[i,7]) # num 6
#  print(lotto_result.iloc[i,8]) # num 7
#  print(lotto_result.iloc[i,9]) # num 7
#  print(lotto_result.iloc[i,10]) # num 7
#  print(lotto_result.iloc[i,11]) # num 7
#  print(lotto_result.iloc[i,12]) # num 7
 
  ws.title = str('Lotto365')
  ws['A'+str(i+2)]=lotto_result.iloc[i,0]
  ws['B'+str(i+2)]=lotto_result.iloc[i,1]

  tmp.append(lotto_result.iloc[i,2])
  tmp.append(lotto_result.iloc[i,3])
  tmp.append(lotto_result.iloc[i,4])
  tmp.append(lotto_result.iloc[i,5])
  tmp.append(lotto_result.iloc[i,6])
  tmp.append(lotto_result.iloc[i,7])
  tmp.sort()
  ws['C'+str(i+2)]=tmp[0]
  ws['D'+str(i+2)]=tmp[1]
  ws['E'+str(i+2)]=tmp[2]
  ws['F'+str(i+2)]=tmp[3]
  ws['G'+str(i+2)]=tmp[4]
  ws['H'+str(i+2)]=tmp[5]
  
  ws['I'+str(i+2)]=lotto_result.iloc[i,8]
  ws['J'+str(i+2)]=lotto_result.iloc[i,9]
  ws['K'+str(i+2)]=lotto_result.iloc[i,10]
  ws['L'+str(i+2)]=lotto_result.iloc[i,11]
  ws['M'+str(i+2)]=lotto_result.iloc[i,12]


  n1  = 0
  n10 = 0
  n20 = 0
  n30 = 0
  n40 = 0

  for j in range(0,6):
    if tmp[j] < 10:
      n1  = n1 + 1
    elif tmp[j] < 20:
      n10 = n10 + 1
    elif tmp[j] < 30:
      n20 = n20 + 1
    elif tmp[j] < 40:
      n30 = n30 + 1
    elif tmp[j] < 46:
      n40 = n40 + 1

  ws['N'+str(i+2)]=n1
  ws['O'+str(i+2)]=n10
  ws['P'+str(i+2)]=n20
  ws['Q'+str(i+2)]=n30
  ws['R'+str(i+2)]=n40
  
  numlist=[1,2,3,4,5,6,7,8,9,10,11,12,13,14,15,16,17,18,19,20,21,22,23,24,25,26,27,28,29,30,31,32,33,34,35,36,37,38,39,40,41,42,43,44,45]
  myindex=0
                                                      # Index : 1734089677 일때
  myindex += (44*43*42*41*40 * numlist.index(tmp[0])) # 130320960씩 빼기 - 1번째 숫자 : 1+ 13회 = 14
  numlist.remove(tmp[0])
  myindex += (43*42*41*40 * numlist.index(tmp[1]))    # 2961840씩 빼기 - 2번째 숫자 : 2+ 13회   = 15
  numlist.remove(tmp[1])
  myindex += (42*41*40 * numlist.index(tmp[2]))       # 68880씩 빼기 - 3번째 숫자 : 3+ 20회     = 23
  numlist.remove(tmp[2])
  myindex += (41*40 * numlist.index(tmp[3]))          # 1640씩 빼기 - 4번째 숫자 : 4+ 21회      = 25
  numlist.remove(tmp[3])
  myindex += (40 * numlist.index(tmp[4]))             # 40씩 빼기 - 5번째 숫자 : 5+ 30회        = 35
  numlist.remove(tmp[4])
  myindex += numlist.index(tmp[5])                    # 1씩 빼기  - 6번째 숫자: 6+ 37회         = 43
  numlist.remove(tmp[5])
  print(myindex)
  numlist.clear()
  ws['S'+str(i+2)] = myindex
  myarr = []
  Retry = []
  for ri in range(0,5):
    Retry.append(True)

  index_init = int(ws['S'+str(i+2)].value * 3339902476 * 0.23224637852690479426)
  if  index_init > 1250000000:
    index_init =  index_init * 3.33 / 1.23
  elif index_init > 500000000:
    index_init =  index_init * 0.93
  elif index_init > 250000000:
    index_init =  index_init * 0.80
  else:
    index_init =  index_init * 0.72

  indexn = index_init
  IsFixNum=False
  while True:
    myarr.append(int(indexn/130320960)+1)
    indexn = indexn % 130320960
    myarr.append(int(indexn/2961840)+2)
    indexn = indexn % 2961840
    myarr.append(int(indexn/68880)+3)
    indexn = indexn % 68880
    myarr.append(int(indexn/1640)+4)
    indexn = indexn % 1640
    myarr.append(int(indexn/40)+5)
    indexn = indexn % 40
    myarr.append(int(indexn)+6)
    if not IsFixNum:
      myarr.sort()
      break
    if myarr[0] >= myarr[1]: indexn = index_init = index_init+2961840
    else: Retry[0] = False
    if myarr[1] >= myarr[2]: indexn = index_init = index_init+68880
    else: Retry[1] = False
    if myarr[2] >= myarr[3]: indexn = index_init = index_init+1640
    else: Retry[2] = False
    if myarr[3] >= myarr[4]: indexn = index_init = index_init+40
    else: Retry[3] = False
    if myarr[4] >= myarr[5]: indexn = index_init = index_init+1
    else: Retry[4] = False
    Bcnt = 0
    for bi in range(0,5):
      if Retry[bi] is False: Bcnt = Bcnt + 1
    if Bcnt == 5: break
    else:
      for ri in range(0,5):
        Retry[ri] = True
      myarr.clear()
  Retry.clear()
  ws['T'+str(i+2)] = myarr[0]
  ws['U'+str(i+2)] = myarr[1]
  ws['V'+str(i+2)] = myarr[2]
  ws['W'+str(i+2)] = myarr[3]
  ws['X'+str(i+2)] = myarr[4]
  ws['Y'+str(i+2)] = myarr[5]
  if i > 0:
    myarr.clear()
    myarr.append(int(ws['T'+str(i+1)].value))
    myarr.append(int(ws['U'+str(i+1)].value))
    myarr.append(int(ws['V'+str(i+1)].value))
    myarr.append(int(ws['W'+str(i+1)].value))
    myarr.append(int(ws['X'+str(i+1)].value))
    myarr.append(int(ws['Y'+str(i+1)].value))
    diffnum=[]
    diffnum.append(int(ws['C'+str(i+2)].value))
    diffnum.append(int(ws['D'+str(i+2)].value))
    diffnum.append(int(ws['E'+str(i+2)].value))
    diffnum.append(int(ws['F'+str(i+2)].value))
    diffnum.append(int(ws['G'+str(i+2)].value))
    diffnum.append(int(ws['H'+str(i+2)].value))
    rnum = 0
    for ck_i in range(0,6):
      for ck_j in range(0,6):
        if diffnum[ck_i] == myarr[ck_j]:
          rnum = rnum + 1
    if rnum < 3:
      ws['Z'+str(i+1)] = "꽝 (" + str(rnum) + ")"
    elif rnum == 3:
      ws['Z'+str(i+1)] = "5등 (" + str(rnum) + ")"
    elif rnum == 4:
      ws['Z'+str(i+1)] = "4등 (" + str(rnum) + ")"
    elif rnum == 5:
      ws['Z'+str(i+1)] = "3등 (" + str(rnum) + ")"
      for ck_i in range(0,6):
        if myarr[ck_i] == ws['I'+str(i+2)].value:
          ws['Z'+str(i+1)] = "2등 (" + str(rnum) + ")"
          break
    elif rnum == 6:
      ws['Z'+str(i+1)] = "1등 (" + str(rnum) + ")"
    ws['AA'+str(i+1)] = float(ws['S'+str(i+2)].value) / float(ws['S'+str(i+1)].value)
    print(str(ws['AA'+str(i+1)]))
    diffnum.clear()
  myarr.clear()
  
  tmp.clear()
  #print('###################')
################################################
# Save And Close File : Excel
################################################
wb.save(Excel_FileName)
wb.close()
################################################
print('Finished.')

# (44*43*42*41*40 * [idx]) +
# (43*42*41*40 * [idx]) +
# (42*41*40 * [idx]) +
# (41*40 * [idx]) +
# (40 * [idx]) +
# [idx]
