class Myproj:
    #def __new__(self):
    #    print('new')
    #    return super().__new__(self)
    def getLastSequence(self):
        import datetime
        dn=0
        w=0
        today = datetime.datetime.today()
        while w < 6:
            tday = today - datetime.timedelta(days=dn)
            a=int(str(tday.year)[0:2])
            b=int(str(tday.year)[2:5])
            c=int(str(tday.month))
            if c <= 2:
                b = b - 1
                c = c + 12
            d=int(str(tday.day))
            w=int((((21*a/4)+(5*b/4)+(26*(c+1)/10)+d-1)%7))
            dn = dn + 1
        wSeq = 0
        sdate = datetime.date(2002,12,7)
        while sdate.year != tday.year or sdate.month != tday.month or sdate.day != tday.day:
            weak = datetime.timedelta(weeks=wSeq)
            sdate = datetime.date(2002,12,7) + weak
            if today.year == sdate.year and today.month == sdate.month and today.day == sdate.day and today.hour < 20 and today.minute < 50: break
            wSeq = wSeq + 1
        return wSeq

    def getInfoFromExcel(self):
        from openpyxl import Workbook
        from openpyxl.styles import Font, Alignment
        from openpyxl.styles import Border, Side
        from openpyxl.styles import PatternFill, Color
        from openpyxl.styles import NamedStyle
        from openpyxl.comments import Comment
        from openpyxl import load_workbook
        info = []
        wb = load_workbook(self.ExcelFileName)
        ws=wb[wb.sheetnames[0]]
        i = 2
        j = 3
        while ws.cell(row=i, column=j).value is not None:
            tmp = []
            while ws.cell(row=i, column=j).value is not None:
                #######################################
                # TO DO:
                #######################################
                if j == 3 or j == 4 or j == 5 or j == 6 or j == 7 or j == 8 or j == 9:
                    tmp.append(ws.cell(row=i, column=j).value)
                    j += 1
                else:
                    break
                #######################################
            if len(tmp) == 7:
                info.append([tmp[0],tmp[1],tmp[2],tmp[3],tmp[4],tmp[5],tmp[6]])
                tmp.clear()
            i += 1
            j = 3
        wb.close()
        # info[seq][0:7] : n1=0, n2=1...bns=6.
        return info

    def __init__(self):
        self.ExcelFileName = './{}.xlsx'.format('Lotto_List')
        self.s_Seq = 1
        self.e_Seq = self.getLastSequence()
        self.info = self.getInfoFromExcel()
        # info[seq][0:7] : n1=0, n2=1...bns=6.

        self.ckRange = 100
        self.maxSetNum=5
        self.StartSeq=0
        self.tmp = self.createNums(self.ckRange)
        self.futureList = []
        for i in range(len(self.tmp)):
            self.futureList.append(list(set(self.tmp[i])))
            self.futureList[i].sort()
        self.checkDiff(self.futureList)
        super().__init__()


    def checkDiff(self,future):
        gn = 5
        cknums_s = 1
        cknums_e = 6
        TotalCheckNum = 0
        self.grade = []
        for i in range(0,5): self.grade.append(0)
        for i in range(self.StartSeq, len(self.info)):
            checkNum = 0
            for j in range(cknums_s-1,cknums_e):
                ckDup = []
                #print(    "futures Num:{}, {}".format(    len(self.future[i-self.StartSeq]), list(set(self.future[i-self.StartSeq]))    )    )
                if self.info[i][j] in future[i-self.StartSeq-1]:      checkNum += 1
                ckDup.clear()
            check2nd=0
            if self.info[i][6] in future[i-self.StartSeq-1]:         check2nd=1
            #print("{}: n{}\t{} ::::: {} ".format( i, checkNum, self.info[i][0:6],self.future[i-(self.StartSeq)]  ) )
            if checkNum >= 3:
                #if checkNum >= 3:
                    #if i > 900: print("{}: n{}\t{} {} {} {} {} {} ::::: {} {} {} {} {} {} ".format( i, checkNum, self.info[i][0],self.info[i][1],self.info[i][2],self.info[i][3],self.info[i][4],self.info[i][5],self.future[i-self.StartSeq][0],self.future[i-self.StartSeq][1],self.future[i-self.StartSeq][2],self.future[i-self.StartSeq][3],self.future[i-self.StartSeq][4],self.future[i-self.StartSeq][5]  ) )
                #print("{}:{}개".format(i+1,checkNum))
                TotalCheckNum += 1
                if    checkNum == 6: self.grade[4] += 1
                elif checkNum == 5:
                    if check2nd > 0: self.grade[3] += 1
                    else: self.grade[2] += 1
                elif checkNum == 4: self.grade[1] += 1
                elif checkNum == 3: self.grade[0] += 1

            if    checkNum == 6:                                 print("1) seq({})\t{}\t:\t({})\t{}".format(i+1,self.info[i][0:6], len(future[i-self.StartSeq-1]),future[i-self.StartSeq-1]).replace('[','').replace(']',''))
            elif checkNum == 5 and check2nd == 1:  print("2) seq({})\t{}\t:\t({})\t{}".format(i+1,self.info[i][0:6], len(future[i-self.StartSeq-1]),future[i-self.StartSeq-1]).replace('[','').replace(']',''))
            elif checkNum == 5:                                  print("3) seq({})\t{}\t:\t({})\t{}".format(i+1,self.info[i][0:6], len(future[i-self.StartSeq-1]),future[i-self.StartSeq-1]).replace('[','').replace(']',''))
            elif checkNum == 4:                                  print("4) seq({})\t{}\t:\t({})\t{}".format(i+1,self.info[i][0:6], len(future[i-self.StartSeq-1]),future[i-self.StartSeq-1]).replace('[','').replace(']',''))
            elif checkNum == 3:                                  print("5) seq({})\t{}\t:\t({})\t{}".format(i+1,self.info[i][0:6], len(future[i-self.StartSeq-1]),future[i-self.StartSeq-1]).replace('[','').replace(']',''))
            else:                                                            print("X) seq({})\t{}\t:\t({})\t{}".format(i+1,self.info[i][0:6], len(future[i-self.StartSeq-1]),future[i-self.StartSeq-1]).replace('[','').replace(']',''))
        print("Latest Num({}): {} ".format(len(future[-1]),future[-1]))
        print("({}/{}*100):{:0.6f}%\t({}/{}) 5:{}, 4:{}, 3:{}, 2:{}, 1:{}".format(TotalCheckNum, (len(self.info)-self.ckRange), TotalCheckNum/(len(self.info)-self.ckRange)*100.0, TotalCheckNum,len(self.info)-1,self.grade[0],self.grade[1],self.grade[2],self.grade[3],self.grade[4]))


    def cleanUp(self):
        for i in range(self.s_Seq - self.s_Seq, self.e_Seq - self.s_Seq):
            if self.info[i]: self.info[i].clear()
        if self.future: self.future.clear()
        for i in range(0,len(self.future)): self.future[i].clear()
        if self.future: self.future.clear()
        if self.grade: self.grade.clear()

    def getValidHistoryList(self, n, info):
        validHistoryList = []
        for i in range(n,n+1):
            validHistoryDict = {}
            for j in range(0, len(info)):
                try: validHistoryDict[info[j][i]] += 1
                except KeyError: validHistoryDict[info[j][i]] = 1

        #keysList = list(validHistoryDict)
        #keysList = list(dict(  sorted(validHistoryDict.items(), key=(lambda x: x[0]) ,reverse=False)))
        if    n == 0: keysList = list(dict(  sorted(validHistoryDict.items(), key=(lambda x: x[0]) ,reverse=True)))
        elif n == 1: keysList = list(dict(  sorted(validHistoryDict.items(), key=(lambda x: x[0]) ,reverse=True)))
        elif n == 2: keysList = list(dict(  sorted(validHistoryDict.items(), key=(lambda x: x[0]) ,reverse=True)))
        elif n == 3: keysList = list(dict(  sorted(validHistoryDict.items(), key=(lambda x: x[0]) ,reverse=False)))
        elif n == 4: keysList = list(dict(  sorted(validHistoryDict.items(), key=(lambda x: x[0]) ,reverse=False)))
        elif n == 5: keysList = list(dict(  sorted(validHistoryDict.items(), key=(lambda x: x[0]) ,reverse=False)))

        for i in range(0,len(keysList)):
            validHistoryList.append(keysList[i])
            validHistoryList.append(validHistoryDict[keysList[i]])
            validHistoryList.append(0)
        return validHistoryList

    def createNums(self,ckRange):
        nums = []    # return nums
        history=[]
        setNum=0
        # ckRange : 한번 번호 추출할때 참조할 회차의 개수 (default: 100)
        self.StartSeq =  len(self.info) - 1
        if len(self.info) > ckRange:            self.StartSeq =  ckRange

        for i in range(self.StartSeq, len(self.info)):
            nums.append([])
            for j in range(0,6):
                nums[i-(self.StartSeq)].append([])
                for k in range(self.maxSetNum): nums[i-(self.StartSeq)][j].append(0)
            for j in range(0,6):
                history.append(self.getValidHistoryList(j,self.info[i-(self.StartSeq):i]))
                for k in range (  int(len(history[j]) / 3)  ): history[j][k*3+2] = 0
            while True:
                for j in range(0,6):
                    modtime=0.90
                    IsBreake = False
                    while IsBreake == False:
                        for k in range (  int(len(history[j]) / 3)  ):
                            if history[j][k*3+1] >= int((ckRange / int(len(history[j]) / 3)) * modtime): history[j][k*3+2] += 1
                        for k in range (  int(len(history[j]) / 3)  ):
                            if (history[j][k*3+2] + history[j][k*3+1] ) % 100 == 0:
                                if nums[i-(self.StartSeq)][j][setNum] == 0: nums[i-(self.StartSeq)][j][setNum] = history[j][k*3+0]
                                setNum += 1
                                if setNum == self.maxSetNum:
                                    setNum=0
                                    break
                        IsBreake = True
                        for k in range(0,self.maxSetNum):
                            if nums[i-(self.StartSeq)][j][k] == 0: IsBreake = False
                if IsBreake: break
            history.clear()
            ret = []
            for i in range(len(nums)):
                ret.append([])
                for j in range(len(nums[i])):
                    for k in range(len(nums[i][j])):
                        ret[i].append(nums[i][j][k])
        nums.clear()
        return ret



    def __del__(self):
        #self.cleanUp()
        pass

def main():
    import sys
    f=Myproj()
    del f


if __name__ == '__main__':
    main()
